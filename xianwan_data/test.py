# -*- coding: utf-8 -*-
# author：Super.Shen

import pandas as pd
import os
from Func import nian, yue, ri, ri_y, df_sum, df_t
import openpyxl

pd.set_option('expand_frame_repr', False)
pd.set_option('display.max_rows', 1000)
import warnings

warnings.filterwarnings('ignore')

# 读取数据
# df = pd.read_excel('C:\\Users\Administrator\Desktop\\闲玩安卓{}{}统计.xlsx'.format(yue, ri_y), index=False)
# print(df.head())


wb = openpyxl.load_workbook('C:\\Users\Administrator\Desktop\\闲玩安卓{}{}统计.xlsx'.format(yue, ri_y))

ws = wb.active
ws.title = "Data"

wb.create_sheet('Summary', 0)
print(wb.active)

wb.save('C:\\Users\Administrator\Desktop\\闲玩安卓{}{}统计.xlsx'.format(yue, ri_y))

wb = openpyxl.load_workbook('C:\\Users\Administrator\Desktop\\闲玩安卓{}{}统计.xlsx'.format(yue, ri_y))

ws = wb.active
ws.title = "Data"

wb.create_sheet('Summary', 0)